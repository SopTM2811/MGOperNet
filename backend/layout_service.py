"""Servicio para generación de layouts SPEI y envío por correo"""
import os
import logging
from pathlib import Path
from datetime import datetime, timezone
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class LayoutService:
    """Servicio para generar layouts SPEI y enviarlos por correo"""
    
    def __init__(self):
        self.layouts_dir = Path("/tmp/netcash_layouts")
        self.layouts_dir.mkdir(exist_ok=True)
        
        # Configuración SMTP con App Password
        self.smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
        self.smtp_port = int(os.getenv("SMTP_PORT", "587"))
        self.smtp_user = os.getenv("SMTP_USER", "")
        self.smtp_pass = os.getenv("SMTP_PASSWORD", "")
        
        # Email de Tesorería (Toño)
        self.tono_email = os.getenv("TONO_EMAIL_LAYOUT", "dfgalezzo@hotmail.com")
        
        smtp_status = "Configurado" if self.smtp_user and self.smtp_pass else "No configurado"
        logger.info(f"LayoutService inicializado. SMTP: {smtp_status}")
    
    def generar_layout_spei(
        self,
        folio_mbco: str,
        clave_mbcontrol: str,
        beneficiarios: List[Dict[str, Any]]
    ) -> str:
        """
        Genera un archivo Excel con el layout SPEI para los beneficiarios.
        
        Args:
            folio_mbco: Folio de la operación NetCash
            clave_mbcontrol: Clave del sistema MBControl (ej: 18434-138-D-11)
            beneficiarios: Lista de dicts con keys: clabe, titular, monto
        
        Returns:
            str: Ruta del archivo Excel generado
        """
        try:
            # Crear nuevo workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Layout SPEI"
            
            # Configurar encabezados
            headers = ["Clabe", "Titular", "Concepto", "Monto"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Llenar datos de beneficiarios
            for row_num, beneficiario in enumerate(beneficiarios, 2):
                ws.cell(row=row_num, column=1).value = beneficiario.get("clabe", "")
                ws.cell(row=row_num, column=2).value = beneficiario.get("titular", "")
                
                # Concepto: PAGO NETCASH [FOLIO_MBco] CLAVE [CLAVE_MBControl]
                concepto = f"PAGO NETCASH {folio_mbco} CLAVE {clave_mbcontrol}"
                ws.cell(row=row_num, column=3).value = concepto
                
                ws.cell(row=row_num, column=4).value = beneficiario.get("monto", 0)
                ws.cell(row=row_num, column=4).number_format = '"$"#,##0.00'
            
            # Ajustar anchos de columnas
            ws.column_dimensions['A'].width = 20  # Clabe
            ws.column_dimensions['B'].width = 35  # Titular
            ws.column_dimensions['C'].width = 50  # Concepto
            ws.column_dimensions['D'].width = 15  # Monto
            
            # Guardar archivo
            filename = f"layout_spei_{folio_mbco}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = self.layouts_dir / filename
            wb.save(filepath)
            
            logger.info(f"Layout SPEI generado: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generando layout SPEI: {str(e)}")
            raise
    
    def enviar_layout_por_correo(
        self,
        layout_path: str,
        folio_mbco: str,
        clave_mbcontrol: str,
        destinatario: str = None
    ) -> bool:
        """
        Envía el layout SPEI por correo electrónico a Tesorería usando Gmail API.
        
        Args:
            layout_path: Ruta del archivo Excel a enviar
            folio_mbco: Folio de la operación
            clave_mbcontrol: Clave MBControl
            destinatario: Email destino (si None, usa TONO_EMAIL_LAYOUT)
        
        Returns:
            bool: True si se envió correctamente
        """
        # Verificar que gmail_service esté disponible
        if not gmail_service:
            logger.warning("Gmail API no disponible. El layout fue generado pero no se pudo enviar.")
            logger.warning(f"Layout generado en: {layout_path}")
            return False
        
        try:
            import asyncio
            
            destinatario = destinatario or self.tono_email
            asunto = f"Layout SPEI NetCash - Folio {folio_mbco} - Clave {clave_mbcontrol}"
            
            # Cuerpo del mensaje
            cuerpo = f"""Hola Toño,

Te envío el layout SPEI para la operación NetCash:

• Folio MBco: {folio_mbco}
• Clave MBControl: {clave_mbcontrol}
• Fecha de generación: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC

Por favor procesa este layout a la brevedad.

Saludos,
Asistente NetCash MBco
"""
            
            # Enviar usando Gmail API con adjunto
            loop = asyncio.get_event_loop()
            if loop.is_running():
                # Si ya hay un loop, crear una tarea
                import concurrent.futures
                with concurrent.futures.ThreadPoolExecutor() as pool:
                    result = pool.submit(
                        asyncio.run,
                        gmail_service.enviar_correo_con_adjuntos(
                            destinatario=destinatario,
                            asunto=asunto,
                            cuerpo=cuerpo,
                            adjuntos=[layout_path]
                        )
                    ).result()
            else:
                result = asyncio.run(
                    gmail_service.enviar_correo_con_adjuntos(
                        destinatario=destinatario,
                        asunto=asunto,
                        cuerpo=cuerpo,
                        adjuntos=[layout_path]
                    )
                )
            
            if result:
                logger.info(f"Layout enviado a {destinatario} via Gmail API")
                return True
            else:
                logger.error("Gmail API retornó None al enviar el layout")
                return False
            
        except Exception as e:
            logger.error(f"Error enviando layout por correo: {str(e)}")
            return False


# Instancia global
layout_service = LayoutService()
